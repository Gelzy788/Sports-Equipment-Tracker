from flask import Flask, render_template, flash, request, redirect, url_for, jsonify, send_file
from flask_login import login_required, login_user, logout_user, current_user
from models import *
from config import app, login_manager, ALLOWED_EXTENSIONS
from database import add_user, update_profile_picture
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from user_app import user_bp
import os
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO

# Проверяем наличие файла базы данных
if not os.path.exists('database.db'):
    print("База данных не найдена. Создаем новую базу данных...")
    from database import init_database
    init_database()
    print("База данных успешно создана!")

login_manager.init_app(app)
login_manager.login_view = 'login'

# Регистрируем Blueprint для пользовательских маршрутов
app.register_blueprint(user_bp, url_prefix='/user')

# Добавляем глобальную переменную для уведомлений


@app.context_processor
def utility_processor():
    def get_notification_count():
        if not current_user.is_authenticated:
            return 0

        if current_user.admin:
            # Для админа - количество новых заявок
            return Requests.query.filter_by(status=None).count()
        else:
            # Для обычного пользователя - количество непрочитанных уведомлений
            return Requests.query.filter(
                Requests.user_id == current_user.id,
                Requests.status_viewed == False,
                Requests.status.in_([0, 1])  # 0 - отклонено, 1 - одобрено
            ).count()

    def get_request_count():
        if current_user.is_authenticated and current_user.admin:
            # Для админа - количество необработанных заявок
            return Requests.query.filter_by(status=None, status_viewed=False).count()
        return 0

    return dict(
        notification_count=get_notification_count(),
        request_count=get_request_count()
    )


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Войти в аккаунт
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # Получите данные из формы
        email = request.form['email']
        password = request.form['password']

        # Найдите пользователя в базе данных
        user = User.query.filter_by(email=email).first()

        # Проверьте пароль
        if user and check_password_hash(user.password, password):
            login_user(user)
            flash("Вы успешно вошли в систему!")
            return redirect(url_for("profile"))
        else:
            flash("Неправильный email или пароль.")

    return render_template("login.html")


# Регистрация пользователя
@app.route("/registration", methods=["GET", "POST"])
def registration():
    if request.method == "POST":
        # Получите данные из формы
        email = request.form['email']
        password = request.form['password']
        username = request.form['username']

        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            flash("Пользователь с таким email уже существует.")
            return redirect(url_for('registration'))

        add_user(email, password, username)
        return redirect(url_for('login'))

    return render_template('registration.html')


# Выйти из аккаунта
@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Вы вышли из системы.")
    return redirect(url_for("get_data"))


# Загрузить изображение
@app.route('/upload', methods=['POST'])
@login_required
def upload_image():
    if 'file' not in request.files:
        flash('Нет файла для загрузки')
        return redirect(url_for('profile'))

    file = request.files['file']

    if file.filename == '':
        flash('Файл не выбран')
        return redirect(url_for('profile'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        current_user.profile_picture = filename
        db.session.commit()
        flash('Файл успешно загружен')
        return redirect(url_for('profile'))

    flash('Недопустимый файл')
    return redirect(url_for('profile'))


# страница профиля
@app.route("/profile")
@login_required
def profile():
    # Получите текущего пользователя
    user = current_user
    return render_template("profile.html", user=user)


# Главная страница
@app.route("/")
def get_data():
    return render_template('index.html')


# Страница инвентаря (склада)
@app.route("/storage", methods=['GET', 'POST'])
@login_required
def storage():
    if not current_user.admin:
        flash('Доступ запрещен')
        return redirect(url_for('index'))

    storage_items = Storage.query.all()
    users = User.query.filter_by(admin=False).all()

    if request.method == 'POST':
        if 'action' in request.form and request.form['action'] == 'give':
            # Логика выдачи инвентаря
            storage_id = request.form.get('storage_id')
            user_id = request.form.get('user_id')
            quantity = int(request.form.get('quantity'))

            storage_item = Storage.query.get(storage_id)
            user = User.query.get(user_id)

            if storage_item and user and storage_item.count >= quantity:
                storage_item.count -= quantity

                existing_equipment = Equipment.query.filter_by(
                    equipment_id=storage_item.id, user_id=user.id).first()

                if existing_equipment:
                    existing_equipment.count += quantity
                else:
                    new_equipment = Equipment(
                        equipment_id=storage_item.id, user_id=user.id, count=quantity)
                    db.session.add(new_equipment)

                db.session.commit()
                flash('Инвентарь успешно передан')
            else:
                flash('Ошибка при передаче инвентаря')

        elif 'action' in request.form and request.form['action'] == 'take_back':
            # Логика забирания инвентаря
            user_id = request.form.get('user_id')
            equipment_id = request.form.get('equipment_id')
            quantity = int(request.form.get('quantity'))

            user = User.query.get(user_id)
            equipment = Equipment.query.filter_by(
                user_id=user.id, equipment_id=equipment_id).first()

            if equipment and equipment.count >= quantity:
                equipment.count -= quantity
                if equipment.count == 0:
                    db.session.delete(equipment)

                # Добавляем количество на склад
                storage_item = Storage.query.get(equipment_id)
                if storage_item:
                    storage_item.count += quantity

                db.session.commit()
                flash('Инвентарь успешно забран у пользователя')
            else:
                flash('Ошибка: недостаточно инвентаря для забора')

    return render_template('storage.html', items=storage_items, users=users)


# Добавление нового предмета
@app.route("/storage/add", methods=['GET', 'POST'])
@login_required
def add_item():
    if not current_user.admin:
        flash("Доступ запрещен")
        return redirect(url_for('get_data'))

    if request.method == 'POST':
        name = request.form.get('name')
        count = request.form.get('count', type=int)

        if name and count is not None:
            # Проверяем, существует ли предмет
            existing_item = Storage.query.filter_by(name=name).first()
            if existing_item:
                # Если предмет существует, увеличиваем его количество
                existing_item.count += count
                flash(f'Количество предмета "{name}" увеличено на {count}')
            else:
                # Если предмет не существует, добавляем его как новый
                new_item = Storage(name=name, count=count)
                db.session.add(new_item)
                flash(f'Предмет "{name}" успешно добавлен')

            db.session.commit()
            return redirect(url_for('storage'))

    return render_template('add_item.html')


# Редактирование предмета
@app.route("/storage/edit/<int:item_id>", methods=['GET', 'POST'])
@login_required
def edit_item(item_id):
    if not current_user.admin:
        flash('У вас нет доступа к этой функции', 'danger')
        return redirect(url_for('storage'))

    item = Storage.query.get_or_404(item_id)

    if request.method == 'POST':
        new_name = request.form.get('name')
        new_count = int(request.form.get('count'))
        new_status = request.form.get('status')

        # Если статус изменился, ищем другие предметы с таким же именем и новым статусом
        if new_status != item.status:
            same_item = Storage.query.filter_by(
                # Убираем "(ремонт)" из имени при поиске
                name=new_name.replace(' (ремонт)', ''),
                status=new_status
            ).first()

            if same_item and same_item.id != item.id:
                # Объединяем с существующим предметом
                same_item.count += new_count
                db.session.delete(item)
                flash('Предмет объединен с существующим', 'success')
            else:
                # Обновляем текущий предмет
                # Убираем "(ремонт)" из имени
                item.name = new_name.replace(' (ремонт)', '')
                item.count = new_count
                item.status = new_status
                flash('Предмет успешно обновлен', 'success')
        else:
            # Просто обновляем данные
            # Убираем "(ремонт)" из имени
            item.name = new_name.replace(' (ремонт)', '')
            item.count = new_count
            item.status = new_status
            flash('Предмет успешно обновлен', 'success')

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении предмета: {str(e)}', 'danger')

        return redirect(url_for('storage'))

    return render_template('edit_item.html', item=item)


# Удаление предмета
@app.route("/storage/delete/<int:item_id>", methods=['DELETE'])
@login_required
def delete_item(item_id):
    if not current_user.admin:
        return jsonify({'error': 'Доступ запрещен'}), 403

    item = Storage.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    return jsonify({'message': 'Предмет удален'}), 200


# Страница управления пользователями
@app.route("/users")
@login_required
def users():
    if not current_user.admin:
        flash("Доступ запрещен")
        return redirect(url_for('get_data'))

    users_list = User.query.all()
    return render_template("users.html", users=users_list)


# Изменение статуса администратора
@app.route("/users/toggle_admin/<int:user_id>", methods=['POST'])
@login_required
def toggle_admin(user_id):
    if not current_user.admin:
        return jsonify({'error': 'Доступ запрещен'}), 403

    # Предотвращаем снятие прав у самого себя
    if user_id == current_user.id:
        return jsonify({'error': 'Нельзя изменить права администратора у самого себя'}), 400

    user = User.query.get_or_404(user_id)
    user.admin = not user.admin
    db.session.commit()

    return jsonify({
        'message': f"Права администратора {'предоставлены' if user.admin else 'отозваны'}",
        'admin': user.admin
    })


# Страница планирования закупок
@app.route("/purchases", methods=['GET', 'POST'])
@login_required
def purchases():
    if not current_user.admin:
        flash("Доступ запрещен")
        return redirect(url_for('get_data'))

    if request.method == 'POST':
        name_eq = request.form.get('name_eq')
        price = request.form.get('price', type=int)
        count = request.form.get('count', type=int)
        supplier = request.form.get('supplier')
        status = request.form.get('status') == '1'

        if name_eq and price is not None and count is not None and supplier:
            new_purchase = Purchases(
                name_eq=name_eq, price=price, count=count, supplier=supplier, status=status)
            db.session.add(new_purchase)

            # Если статус "выполнено", добавляем товар на склад
            if status:
                item = Storage.query.filter_by(name=name_eq).first()
                if item:
                    item.count += count
                else:
                    new_item = Storage(name=name_eq, count=count)
                    db.session.add(new_item)

            db.session.commit()
            flash('План закупки успешно добавлен')
            return redirect(url_for('purchases'))

    purchase_plans = Purchases.query.all()
    return render_template("purchases.html", purchases=purchase_plans)


# Изменение статуса закупки
@app.route("/purchases/toggle_status/<int:purchase_id>", methods=['POST'])
@login_required
def toggle_purchase_status(purchase_id):
    if not current_user.admin:
        return jsonify({'error': 'Доступ запрещен'}), 403

    purchase = Purchases.query.get_or_404(purchase_id)
    item = Storage.query.filter_by(name=purchase.name_eq).first()

    if purchase.status:  # Если статус "выполнено", пытаемся отменить
        if item and item.count >= purchase.count:
            item.count -= purchase.count
            purchase.status = False
            db.session.commit()
            return jsonify({
                'message': f"Статус закупки изменен на 'не выполнено'. Количество предметов '{purchase.name_eq}' уменьшено на складе.",
                'status': purchase.status
            })
        else:
            return jsonify({'error': f"Недостаточно предметов '{purchase.name_eq}' на складе для отмены выполнения закупки."}), 400
    else:  # Если статус "не выполнено", пытаемся выполнить
        if item:
            item.count += purchase.count
        else:
            new_item = Storage(name=purchase.name_eq, count=purchase.count)
            db.session.add(new_item)

        purchase.status = True
        db.session.commit()
        return jsonify({
            'message': f"Статус закупки изменен на 'выполнено'. Предметы '{purchase.name_eq}' добавлены на склад.",
            'status': purchase.status
        })


# Удаление закупки
@app.route("/purchases/delete/<int:purchase_id>", methods=['DELETE'])
@login_required
def delete_purchase(purchase_id):
    if not current_user.admin:
        return jsonify({'error': 'Доступ запрещен'}), 403

    purchase = Purchases.query.get_or_404(purchase_id)
    db.session.delete(purchase)
    db.session.commit()

    return jsonify({'message': 'Закупка удалена'}), 200


@app.route('/my_inventory')
@login_required
def my_inventory():
    equipment = Equipment.query.filter_by(user_id=current_user.id).all()
    return render_template('my_inventory.html', equipment=equipment)


@app.route("/user_inventory/<int:user_id>")
@login_required
def user_inventory(user_id):
    user_equipment = Equipment.query.filter_by(user_id=user_id).all()
    return jsonify({
        'equipment': [{
            'equipment_id': item.equipment_id,
            'storage': Storage.query.get(item.equipment_id),
            'count': item.count
        } for item in user_equipment]
    })


@app.route("/profile/<int:user_id>")
@login_required
def user_profile(user_id):
    user = User.query.get_or_404(user_id)
    user_equipment = Equipment.query.filter_by(user_id=user.id).all()
    return render_template("user_profile.html", user=user, equipment=user_equipment)


# Страница заявок (для всех пользователей)
@app.route("/requests")
@login_required
def requests():
    if current_user.admin:
        # Для админа показываем все заявки с возможностью управления
        requests = Requests.query.order_by(Requests.created_at.desc()).all()
        return render_template('admin_requests.html', requests=requests)
    else:
        # Для обычного пользователя показываем только его заявки
        user_requests = Requests.query.filter_by(
            user_id=current_user.id).order_by(Requests.created_at.desc()).all()
        return render_template('my_requests.html', requests=user_requests)


# Обработка заявки (одобрение/отклонение)
@app.route("/admin/request/<int:request_id>/<action>", methods=['POST'])
@login_required
def process_request(request_id, action):
    try:
        request = Requests.query.get_or_404(request_id)

        if action == 'approve':
            if request.equipment_id:  # Для обычных заявок
                storage_item = Storage.query.get(request.equipment_id)

                if not storage_item:
                    flash('Ошибка: предмет не найден на складе', 'danger')
                    return redirect(url_for('requests'))

                if request.request_type == 'ремонт':
                    # Находим оборудование пользователя
                    user_equipment = Equipment.query.filter_by(
                        user_id=request.user_id,
                        equipment_id=request.equipment_id
                    ).first()

                    if user_equipment:
                        # Перемещаем оборудование на склад со статусом "ремонт"
                        broken_storage = Storage.query.filter_by(
                            name=storage_item.name,
                            status='ремонт'
                        ).first()

                        if broken_storage:
                            # Если уже есть оборудование в ремонте этого типа
                            broken_storage.count += request.count
                        else:
                            # Создаем новую запись для оборудования в ремонте
                            broken_storage = Storage(
                                name=storage_item.name,
                                count=request.count,
                                status='ремонт'
                            )
                            db.session.add(broken_storage)

                        # Удаляем оборудование у пользователя
                        if user_equipment.count > request.count:
                            user_equipment.count -= request.count
                        else:
                            db.session.delete(user_equipment)

                    request.status = True
                    request.status_viewed = False
                    request.status_changed_at = datetime.utcnow()
                    flash(
                        'Заявка на ремонт одобрена, оборудование перемещено на склад для ремонта', 'success')
                else:
                    # Проверяем все активные заявки на это оборудование
                    active_requests = Requests.query.filter_by(
                        equipment_id=request.equipment_id,
                        status=None
                    ).order_by(Requests.created_at).all()

                    # Считаем общее количество запрошенного оборудования в активных заявках
                    total_requested = sum(
                        r.count for r in active_requests if r.id <= request.id)

                    if storage_item.count >= total_requested:
                        request.status = True
                        request.status_viewed = False
                        request.status_changed_at = datetime.utcnow()
                        storage_item.count -= request.count

                        # Проверяем, есть ли уже такое оборудование у пользователя
                        user_equipment = Equipment.query.filter_by(
                            user_id=request.user_id,
                            equipment_id=request.equipment_id
                        ).first()

                        if user_equipment:
                            # Если есть - увеличиваем количество
                            user_equipment.count += request.count
                        else:
                            # Если нет - создаем новую запись
                            new_equipment = Equipment(
                                user_id=request.user_id,
                                equipment_id=request.equipment_id,
                                count=request.count
                            )
                            db.session.add(new_equipment)

                        flash('Заявка одобрена', 'success')
                    else:
                        flash(
                            'Недостаточно оборудования на складе с учетом всех активных заявок', 'danger')
                        return redirect(url_for('requests'))
            else:  # Для кастомных заявок
                # Проверяем, существует ли уже такой предмет на складе
                existing_item = Storage.query.filter_by(
                    name=request.equipment_name).first()

                if not existing_item:
                    flash(
                        'Ошибка: Нельзя одобрить кастомную заявку. Сначала добавьте предмет на склад.', 'danger')
                    return redirect(url_for('requests'))

                if existing_item.count < (request.count or 1):
                    flash('Ошибка: Недостаточно предметов на складе', 'danger')
                    return redirect(url_for('requests'))

                try:
                    # Обновляем заявку с id существующего предмета
                    request.equipment_id = existing_item.id
                    request.status = True
                    request.status_viewed = False
                    request.status_changed_at = datetime.utcnow()

                    # Уменьшаем количество на складе
                    existing_item.count -= (request.count or 1)

                    # Создаем запись о выдаче оборудования пользователю
                    new_equipment = Equipment(
                        user_id=request.user_id,
                        equipment_id=existing_item.id,
                        count=request.count or 1
                    )
                    db.session.add(new_equipment)
                    flash('Кастомная заявка одобрена', 'success')
                except Exception as e:
                    print(f"Ошибка при обработке кастомной заявки: {str(e)}")
                    flash('Ошибка при обработке кастомной заявки', 'danger')
                    return redirect(url_for('requests'))

        elif action == 'reject':
            request.status = False
            request.status_viewed = False
            request.status_changed_at = datetime.utcnow()
            flash('Заявка отклонена', 'danger')

        db.session.commit()
        return redirect(url_for('requests'))

    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при обработке заявки: {str(e)}', 'danger')
        return redirect(url_for('requests'))


@app.route('/api/equipment/search')
@login_required
def search_equipment():
    term = request.args.get('term', '')
    if len(term) < 2:
        return jsonify([])

    # Поиск оборудования по названию
    equipment = Storage.query.filter(Storage.name.ilike(f'%{term}%')).all()
    return jsonify([{
        'id': item.id,
        'name': item.name,
        'description': item.description
    } for item in equipment])


@app.route('/create_custom_request', methods=['POST'])
@login_required
def create_custom_request():
    equipment_name = request.form.get('equipment_name')
    custom_description = request.form.get('custom_description')
    request_type = request.form.get('request_type')

    # Поиск оборудования по названию
    equipment = Storage.query.filter(Storage.name == equipment_name).first()

    # Создаем новый запрос
    new_request = Requests(
        user_id=current_user.id,
        storage_id=equipment.id if equipment else None,
        request_type=request_type,
        custom_description=custom_description
    )

    db.session.add(new_request)
    db.session.commit()

    flash('Ваш запрос успешно создан и отправлен на рассмотрение', 'success')
    return redirect(url_for('requests'))


@app.route('/mark_requests_viewed', methods=['POST'])
@login_required
def mark_requests_viewed():
    if not current_user.admin:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403

    try:
        # Получаем все необработанные заявки
        unprocessed_requests = Requests.query.filter_by(status=None).all()

        # Отмечаем их как просмотренные
        for request in unprocessed_requests:
            request.status_viewed = True

        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route("/export_inventory")
@login_required
def export_inventory():
    if not current_user.admin:
        flash('Доступ запрещен')
        return redirect(url_for('storage'))

    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Инвентарь"

    # Добавляем заголовки
    headers = ['Имя пользователя', 'Предмет', 'Количество', 'Статус']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Получаем данные
    row = 2
    users = User.query.all()
    for user in users:
        equipment = Equipment.query.filter_by(user_id=user.id).all()
        for item in equipment:
            storage_item = Storage.query.get(item.equipment_id)
            ws.cell(row=row, column=1, value=user.username)
            ws.cell(row=row, column=2, value=storage_item.name)
            ws.cell(row=row, column=3, value=item.count)
            ws.cell(row=row, column=4, value=storage_item.status)
            row += 1

    # Настраиваем ширину столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем в BytesIO объект
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='inventory_report.xlsx'
    )


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
